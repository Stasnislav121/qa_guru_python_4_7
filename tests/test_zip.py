import csv
import os
import pathlib
from zipfile import ZipFile

import pytest
from PyPDF2 import PdfReader
from openpyxl import load_workbook

path = '../examples_files'
directory = pathlib.Path(path)


@pytest.fixture(autouse=True)
def create_zip():
    file_size_create = []
    with ZipFile("../resources/Архив.zip", "w") as myzip:
        for file_path in directory.iterdir():
            print(file_path.name)
            myzip.write(file_path, arcname=file_path.name)
            file_size_create.append(os.path.getsize(file_path))
    yield file_size_create


@pytest.fixture()
def row_pdf():
    with open(f"{directory}/docs-pytest-org-en-latest.pdf", "rb") as pdf_1:
        reader = PdfReader(pdf_1)
        numbers_of_pages_before = len(reader.pages)
        text_before = reader.pages[numbers_of_pages_before - 1].extract_text()
    yield numbers_of_pages_before, text_before


@pytest.fixture()
def row_xlsx():
    workbook = load_workbook(f"{directory}/file_example_XLSX_10.xlsx")
    sheet = workbook.active
    xlsx_rows_before = sheet.max_row
    xlsx_cell_before = sheet.cell(row=3, column=2).value
    yield xlsx_rows_before, xlsx_cell_before


@pytest.fixture()
def count_row_csv():
    with open(f"{directory}/username.csv") as csvfile_2:
        csvfile2 = csv.reader(csvfile_2)
        count_row_before = len(list(csvfile2))
    with open(f"{directory}/username.csv") as csvfile:
        csvfile2 = csv.reader(csvfile)
        for row in csvfile2:
            last_row_before = row
    yield count_row_before, last_row_before


def test_file_availability(create_zip):
    file_weight_before_creation = create_zip
    file_weight_after_creation = []
    with ZipFile("../resources/Архив.zip") as zip_:
        assert zip_.getinfo("docs-pytest-org-en-latest.pdf")
        assert zip_.getinfo("file_example_XLSX_10.xlsx")
        assert zip_.getinfo("username.csv")
        for item in zip_.infolist():
            file_weight_after_creation.append(item.file_size)
    assert file_weight_after_creation == file_weight_before_creation


def test_check_pdf(create_zip, row_pdf):
    numbers_of_pages_before = row_pdf[0]
    text_before = row_pdf[1]
    with ZipFile("../resources/Архив.zip") as zip_:
        with zip_.open("docs-pytest-org-en-latest.pdf") as f_pdf:
            reader = PdfReader(f_pdf)
            numbers_of_pages_after = len(reader.pages)
            text_after = reader.pages[numbers_of_pages_after - 1].extract_text()
    assert numbers_of_pages_before == numbers_of_pages_after
    assert text_before == text_after


def test_check_xlsx(create_zip, row_xlsx):
    xlsx_rows_before = row_xlsx[0]
    xlsx_cell_before = row_xlsx[1]
    with ZipFile("../resources/Архив.zip") as zip_:
        with zip_.open("file_example_XLSX_10.xlsx") as f_xlsx:
            workbook = load_workbook(f_xlsx)
            sheet = workbook.active
            xlsx_rows_after = sheet.max_row
            xlsx_cell_after = sheet.cell(row=3, column=2).value
    assert xlsx_rows_before == xlsx_rows_after
    assert xlsx_cell_before == xlsx_cell_after


def test_check_csv(create_zip, count_row_csv):
    count_row_before = count_row_csv[0]
    last_row_before = count_row_csv[1][0]
    with ZipFile("../resources/Архив.zip") as zip_:
        with zip_.open("username.csv") as f_csv:
            count_row_after = len(list(f_csv))
        with zip_.open("username.csv") as csvfile:
            for row in csvfile:
                last_row_after = row
        last_row_after = last_row_after.decode("utf-8", "ignore")
    assert count_row_before == count_row_after
    assert last_row_before in last_row_after
